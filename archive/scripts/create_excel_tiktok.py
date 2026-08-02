import os
import sys
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
    
    excel_file = "trending_affiliate_products.xlsx"
    print(f"Creating highly styled TikTok Excel file: {excel_file}...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TikTok Affiliate Products"
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Define styles
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="FE2C55") # TikTok Red Color
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="333333")
    link_font = Font(name=font_family, size=10, color="00B6C2", underline="single") # TikTok Teal Link
    
    header_fill = PatternFill(start_color="121212", end_color="121212", fill_type="solid") # Dark / Black Header
    zebra_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid") # Very light grey Zebra
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E5E5'),
        right=Side(style='thin', color='E5E5E5'),
        top=Side(style='thin', color='E5E5E5'),
        bottom=Side(style='thin', color='E5E5E5')
    )
    
    # Add title row
    ws.merge_cells("A1:I1")
    ws["A1"] = "สินค้าขายดีอันดับยอดนิยมสำหรับทำวิดีโอ AI (TikTok Shop Affiliate Thailand)"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Add an empty row
    ws.row_dimensions[2].height = 15
    
    # Add headers
    headers = [
        "อันดับ (Rank)",
        "รหัสสินค้า (Product ID)",
        "ชื่อสินค้า (Product Name)",
        "หมวดหมู่ (Category)",
        "ราคาขาย (Price)",
        "ค่าคอมมิชชัน (%)",
        "รายได้ต่อชิ้น (Est. Earnings)",
        "ลิงก์สินค้า (Product URL)",
        "แนวคิด/บทพูดสั้น (AI Video Concept)"
    ]
    
    ws.append([]) # dummy row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[3].height = 30
    
    # TikTok Product Data
    products = [
        {
            "rank": 1, "id": "1729456727501474015", "name": "PAPA FEEL 577 เซรั่มลดเลือนฝ้ากระและรอยสิว",
            "cat": "Beauty & Personal Care", "price": 250.00, "comm": 0.20, "earn": 50.00,
            "url": "https://shop.tiktok.com/view/product/1729456727501474015?region=TH&locale=th-TH",
            "concept": "โชว์แก้มก่อนทาเซรั่มที่มีฝ้ากระชัดเจน แล้วทาทับ เผยภาพผิวหน้าใสเนียนกริบ (ผลลัพธ์ใน 3 วินาทีแรก)"
        },
        {
            "rank": 2, "id": "1730678936346200804", "name": "เครื่องนวดหน้ากัวซาไฟฟ้า FULI (FULI Electric Guasha)",
            "cat": "Beauty & Personal Care", "price": 890.00, "comm": 0.15, "earn": 133.50,
            "url": "https://shop.tiktok.com/view/product/1730678936346200804?region=TH&locale=th-TH",
            "concept": "คนตื่นนอนหน้าบวมเป่ง แล้วใช้เครื่องกัวซานวดหน้าเพียง 1 นาที หน้าเรียวกระชับขึ้นทันที (ก่อน-หลัง)"
        },
        {
            "rank": 3, "id": "1731861070669187858", "name": "Beleaf Liposomal Vitamin C (วิตามินซีเข้มข้น)",
            "cat": "Food & Supplements", "price": 390.00, "comm": 0.15, "earn": 58.50,
            "url": "https://shop.tiktok.com/view/product/1731861070669187858?region=TH&locale=th-TH",
            "concept": "ไขข้อข้องใจ 'ทำไมกินวิตามินซีปกติแล้วผิวไม่ใสสักที?' เผยเทคโนโลยีไลโปโซมอลที่ดูดซึมดีกว่า 10 เท่า"
        },
        {
            "rank": 4, "id": "1731576750018628290", "name": "กันแดดเนื้อน้ำนม Hanbyeol (Hanbyeol Sunscreen Milk)",
            "cat": "Beauty & Personal Care", "price": 290.00, "comm": 0.20, "earn": 58.00,
            "url": "https://shop.tiktok.com/view/product/1731576750018628290?region=TH&locale=th-TH",
            "concept": "ทากันแดดแล้วเอาทิชชูซับหน้าให้ดูว่าไม่มีความมันเหนอะหนะ โชว์ความเป็นธรรมชาติ ท้าแดดเมืองไทย"
        },
        {
            "rank": 5, "id": "1731527642712148180", "name": "Smooth Bio C 1000 mg (วิตามินซีแบบเคี้ยว)",
            "cat": "Food & Supplements", "price": 150.00, "comm": 0.15, "earn": 22.50,
            "url": "https://www.tiktok.com/view/product/1731527642712148180?region=TH&locale=th-TH",
            "concept": "เผลอกินขนมหวานเยอะจนเจ็บคอ ร่างกายอ่อนแอ ลองวิตามินซีเคี้ยวอร่อย ได้สุขภาพดีในราคาหลักร้อย"
        },
        {
            "rank": 6, "id": "1732374653152432971", "name": "น้ำยาบ้วนปาก Buoola (Buoola Mouthwash)",
            "cat": "Beauty & Personal Care", "price": 180.00, "comm": 0.15, "earn": 27.00,
            "url": "https://www.tiktok.com/view/product/1732374653152432971?region=TH&locale=th-TH",
            "concept": "ปัญหากลิ่นปากแรงหลังกินส้มตำ/กระเทียม บ้วน Buoola เพียง 30 วินาที คราบหลุดออกหมด ลมหายใจหอมสดชื่น"
        },
        {
            "rank": 7, "id": "1731667000536500824", "name": "เกล็ดขนมปัง ตราครัววังทิพย์ 200 กรัม",
            "cat": "Food & Kitchen", "price": 45.00, "comm": 0.10, "earn": 4.50,
            "url": "https://shop.tiktok.com/view/product/1731667000536500824?region=TH&locale=th-TH",
            "concept": "ทำไก่ทอดกรอบสีเหลืองทองกรุบกรอบน่ากินที่บ้านง่ายๆ ด้วยเกล็ดขนมปังเกรดพรีเมียม"
        },
        {
            "rank": 8, "id": "1731283808641845848", "name": "ชามพลาสติก PP213 จำนวน 50 ชิ้น",
            "cat": "Kitchen & Home", "price": 80.00, "comm": 0.10, "earn": 8.00,
            "url": "https://shop.tiktok.com/view/product/1731283808641845848?region=TH&locale=th-TH",
            "concept": "จัดระเบียบการแพ็คอาหารส่งเดลิเวอรี่ให้สวยงามด้วยชามฝาล็อคแน่นหนา น้ำซุปไม่มีหกเลอะเทอะ"
        },
        {
            "rank": 9, "id": "1731860742826396434", "name": "Alpha Chlorophyll Plus (อาหารเสริมคลอโรฟิลล์)",
            "cat": "Food & Supplements", "price": 290.00, "comm": 0.15, "earn": 43.50,
            "url": "https://shop.tiktok.com/view/product/1731860742826396434?region=TH&locale=th-TH",
            "concept": "หน้าท้องป่อง ขับถ่ายยาก ชงดีท็อกซ์คลอโรฟิลล์ก่อนนอน ตื่นเช้ามาพุงยุบ สบายตัวสุดๆ"
        },
        {
            "rank": 10, "id": "1731194000810019602", "name": "วิตามินบำรุงผิว Colla C (Colla C Collagen)",
            "cat": "Food & Supplements", "price": 350.00, "comm": 0.18, "earn": 63.00,
            "url": "https://shop.tiktok.com/view/product/1731194000810019602?region=TH&locale=th-TH",
            "concept": "เคล็ดลับผิวขาวใสออร่าท้าแดด ชงคอลลาเจนกินง่าย รสชาติอร่อย ผิวนุ่มลื่นใน 7 วัน"
        }
    ]
    
    current_row = 4
    for p in products:
        ws.cell(row=current_row, column=1, value=p["rank"])
        ws.cell(row=current_row, column=2, value=p["id"])
        ws.cell(row=current_row, column=3, value=p["name"])
        ws.cell(row=current_row, column=4, value=p["cat"])
        ws.cell(row=current_row, column=5, value=p["price"])
        ws.cell(row=current_row, column=6, value=p["comm"])
        ws.cell(row=current_row, column=7, value=p["earn"])
        
        # Hyperlink
        url_cell = ws.cell(row=current_row, column=8, value="Open Link")
        url_cell.hyperlink = p["url"]
        url_cell.font = link_font
        
        ws.cell(row=current_row, column=9, value=p["concept"])
        
        is_even = (current_row % 2 == 0)
        for col_idx in range(1, 10):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            if col_idx != 8:
                cell.font = data_font
            
            if col_idx in [1, 2]:
                cell.alignment = align_center
            elif col_idx in [3, 4, 9]:
                cell.alignment = align_left
            elif col_idx in [5, 6, 7]:
                cell.alignment = align_right
            elif col_idx == 8:
                cell.alignment = align_center
                
            if col_idx in [5, 7]:
                cell.number_format = '฿#,##0.00'
            elif col_idx == 6:
                cell.number_format = '0.0%'
                
            if is_even:
                cell.fill = zebra_fill
                
        ws.row_dimensions[current_row].height = 45
        current_row += 1
        
    # Auto-adjust column widths
    column_widths = {
        "A": 15, "B": 25, "C": 40, "D": 25, "E": 18, "F": 18, "G": 20, "H": 18, "I": 75
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    wb.save(excel_file)
    print(f"Successfully updated Excel with TikTok products: {excel_file}")
    
    # Save CSV version also
    csv_file = "trending_affiliate_products.csv"
    with open(csv_file, mode='w', encoding='utf-8-sig', newline='') as f:
        fieldnames = ["Product ID", "Product Name", "Category", "Price (THB)", "Commission Rate", "Product URL", "Google Sheets Link", "AI Video Hook / Concept"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for p in products:
            writer.writerow({
                "Product ID": p["id"],
                "Product Name": p["name"],
                "Category": p["cat"],
                "Price (THB)": f"{p['price']:.2f}",
                "Commission Rate": f"{int(p['comm']*100)}%",
                "Product URL": p["url"],
                "Google Sheets Link": f'=HYPERLINK("{p["url"]}", "Open Link")',
                "AI Video Hook / Concept": p["concept"]
            })
    print(f"Successfully updated CSV with TikTok products: {csv_file}")

if __name__ == "__main__":
    main()
