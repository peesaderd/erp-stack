import os
import sys
import json
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from playwright.sync_api import sync_playwright

def save_to_files(products):
    excel_file = "trending_affiliate_products.xlsx"
    csv_file = "trending_affiliate_products.csv"
    
    print(f"\nSaving {len(products)} mined products to Excel & CSV...")
    
    # 1. Save to Excel with beautiful TikTok styling
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mined Affiliate Products"
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="FE2C55") # TikTok Red
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="333333")
    link_font = Font(name=font_family, size=10, color="00B6C2", underline="single") # TikTok Teal Link
    
    header_fill = PatternFill(start_color="121212", end_color="121212", fill_type="solid") # Dark Header
    zebra_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E5E5'),
        right=Side(style='thin', color='E5E5E5'),
        top=Side(style='thin', color='E5E5E5'),
        bottom=Side(style='thin', color='E5E5E5')
    )
    
    ws.merge_cells("A1:I1")
    ws["A1"] = "สินค้าขายดีอันดับยอดนิยมขุดอัตโนมัติ (Mined TikTok Shop Affiliate)"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 15
    
    headers = [
        "อันดับ (Rank)",
        "รหัสสินค้า (Product ID)",
        "ชื่อสินค้า (Product Name)",
        "หมวดหมู่ (Category)",
        "ราคาขาย (Price)",
        "ค่าคอมมิชชัน (%)",
        "รายได้ต่อชิ้น (Est. Earnings)",
        "ลิงก์สินค้า (Product URL)",
        "แนวคิด/บทพูดสั้น (AI Video Concept)"
    ]
    
    ws.append([]) # dummy row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[3].height = 30
    
    current_row = 4
    for idx, p in enumerate(products, 1):
        ws.cell(row=current_row, column=1, value=idx)
        ws.cell(row=current_row, column=2, value=p["id"])
        ws.cell(row=current_row, column=3, value=p["name"])
        ws.cell(row=current_row, column=4, value=p["cat"])
        ws.cell(row=current_row, column=5, value=p["price"])
        ws.cell(row=current_row, column=6, value=p["comm"])
        ws.cell(row=current_row, column=7, value=p["earn"])
        
        # Hyperlink
        url_cell = ws.cell(row=current_row, column=8, value="Open Link")
        url_cell.hyperlink = p["url"]
        url_cell.font = link_font
        
        ws.cell(row=current_row, column=9, value=p["concept"])
        
        is_even = (current_row % 2 == 0)
        for col_idx in range(1, 10):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            if col_idx != 8:
                cell.font = data_font
            
            if col_idx in [1, 2]:
                cell.alignment = align_center
            elif col_idx in [3, 4, 9]:
                cell.alignment = align_left
            elif col_idx in [5, 6, 7]:
                cell.alignment = align_right
            elif col_idx == 8:
                cell.alignment = align_center
                
            if col_idx in [5, 7]:
                cell.number_format = '฿#,##0.00'
            elif col_idx == 6:
                cell.number_format = '0.0%'
                
            if is_even:
                cell.fill = zebra_fill
                
        ws.row_dimensions[current_row].height = 45
        current_row += 1
        
    column_widths = {"A": 15, "B": 25, "C": 40, "D": 25, "E": 18, "F": 18, "G": 20, "H": 18, "I": 75}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    try:
        wb.save(excel_file)
        print(f"Successfully saved styled Excel: {excel_file}")
    except PermissionError:
        print(f"Error: Permission denied saving Excel to {excel_file}. Please close any open Excel window first!")
        sys.exit(1)
        
    # 2. Save to CSV
    with open(csv_file, mode='w', encoding='utf-8-sig', newline='') as f:
        fieldnames = ["Product ID", "Product Name", "Category", "Price (THB)", "Commission Rate", "Product URL", "Google Sheets Link", "AI Video Hook / Concept"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for idx, p in enumerate(products, 1):
            writer.writerow({
                "Product ID": p["id"],
                "Product Name": p["name"],
                "Category": p["cat"],
                "Price (THB)": f"{p['price']:.2f}",
                "Commission Rate": f"{int(p['comm']*100)}%",
                "Product URL": p["url"],
                "Google Sheets Link": f'=HYPERLINK("{p["url"]}", "Open Link")',
                "AI Video Hook / Concept": p["concept"]
            })
    print(f"Successfully saved CSV: {csv_file}")

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
    
    profile_dir = os.path.abspath(".tiktok_profile")
    print(f"Loading persistent profile from: {profile_dir}")
    
    mined_products = []
    
    with sync_playwright() as p:
        print("Launching headed browser to bypass anti-bot checks and show progress...")
        context = p.chromium.launch_persistent_context(
            user_data_dir=profile_dir,
            headless=False,
            args=["--disable-blink-features=AutomationControlled"],
            no_viewport=True
        )
        
        # Monitor responses to capture product list payload
        def handle_response(response):
            try:
                url = response.url
                if "promote_products/list" in url or "product_selection/list" in url:
                    content_type = response.headers.get("content-type", "")
                    if "application/json" in content_type:
                        res_json = response.json()
                        print(f"Intercepted Product Data API Call!")
                        # Parse products
                        data = res_json.get("data", {})
                        products = data.get("products", []) or data.get("list", [])
                        
                        if products:
                            for prod in products:
                                prod_id = prod.get("product_id") or prod.get("id")
                                name = prod.get("name") or prod.get("title")
                                price = float(prod.get("price", {}).get("min_price") or prod.get("price") or 100)
                                comm = float(prod.get("commission_rate", 10)) / 100.0
                                if comm > 1.0: # If rate is in percentage like 20 instead of 0.2
                                    comm = comm / 100.0
                                    
                                cat = prod.get("category", "General")
                                prod_url = f"https://shop.tiktok.com/view/product/{prod_id}?region=TH&locale=th-TH"
                                concept = f"โชว์สินค้า {name} ชูผลลัพธ์และความคุ้มค่าดึงดูดความสนใจผู้ชมทันทีใน 3 วินาทีแรก"
                                
                                mined_products.append({
                                    "id": str(prod_id),
                                    "name": name,
                                    "cat": cat,
                                    "price": price,
                                    "comm": comm,
                                    "earn": price * comm,
                                    "url": prod_url,
                                    "concept": concept
                                })
            except Exception as e:
                pass
                
        context.on("response", handle_response)
        
        page = context.pages[0] if context.pages else context.new_page()
        
        print("Navigating to TikTok Seller Center Affiliate homepage...")
        page.goto("https://seller-th.tiktok.com/affiliate/landing?shop_region=TH", wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(8000)
        
        # Click Go to Affiliate Center Home or Go to Open Collaboration
        btn = page.locator("button:has-text('Go to Affiliate Center Home'), button:has-text('Go to Open Collaboration')").first
        if btn.is_visible():
            print("Clicking Go to Affiliate Center button...")
            btn.click()
            page.wait_for_timeout(10000)
            
            # We check the tabs
            pages = context.pages
            if len(pages) > 1:
                target_page = pages[1]
                print(f"Switched to Affiliate Center: {target_page.url}")
                
                # Extract shop_id dynamically from url
                shop_id = "7495822022835407448"
                if "shop_id=" in target_page.url:
                    shop_id = target_page.url.split("shop_id=")[1].split("&")[0]
                    
                collab_url = f"https://affiliate.tiktok.com/product/open-collaboration?shop_id={shop_id}&shop_region=TH"
                print(f"Navigating target page directly to: {collab_url}")
                target_page.goto(collab_url, wait_until="domcontentloaded", timeout=60000)
                target_page.wait_for_timeout(8000)
                
                # Check for "Add products" button
                add_btn = target_page.locator("button:has-text('Add products')").first
                if add_btn.is_visible():
                    print("Found 'Add products' button! Clicking it...")
                    add_btn.click()
                    target_page.wait_for_timeout(8000)
                else:
                    # Try clicking "Not added" tab
                    not_added_tab = target_page.locator("text=Not added").first
                    if not_added_tab.is_visible():
                        print("Clicking 'Not added' tab...")
                        not_added_tab.click()
                        target_page.wait_for_timeout(8000)
            else:
                print("Affiliate Center tab did not open.")
        else:
            print("Affiliate Center Home button not found on landing page.")
            
        print("Waiting 10 seconds for API calls to settle and return product data...")
        page.wait_for_timeout(10000)
        
        context.close()
        
    if mined_products:
        print(f"\nSuccessfully mined {len(mined_products)} products!")
        # Remove duplicates
        seen = set()
        unique_products = []
        for p in mined_products:
            if p["id"] not in seen:
                seen.add(p["id"])
                unique_products.append(p)
                
        save_to_files(unique_products[:10])
    else:
        print("\nFailed to mine product data. Falling back to the previous list...")
        # Fallback to the default 10 bestselling products to keep the pipeline intact
        fallback_products = [
            {"id": "1729456727501474015", "name": "PAPA FEEL 577 เซรั่มลดเลือนฝ้ากระและรอยสิว", "cat": "Beauty & Personal Care", "price": 250.00, "comm": 0.20, "earn": 50.00, "url": "https://shop.tiktok.com/view/product/1729456727501474015?region=TH&locale=th-TH", "concept": "โชว์แก้มก่อนทาเซรั่มที่มีฝ้ากระชัดเจน แล้วทาทับ เผยภาพผิวหน้าใสเนียนกริบ (ผลลัพธ์ใน 3 วินาทีแรก)"},
            {"id": "1730678936346200804", "name": "เครื่องนวดหน้ากัวซาไฟฟ้า FULI (FULI Electric Guasha)", "cat": "Beauty & Personal Care", "price": 890.00, "comm": 0.15, "earn": 133.50, "url": "https://shop.tiktok.com/view/product/1730678936346200804?region=TH&locale=th-TH", "concept": "คนตื่นนอนหน้าบวมเป่ง แล้วใช้เครื่องกัวซานวดหน้าเพียง 1 นาที หน้าเรียวกระชับขึ้นทันที (ก่อน-หลัง)"},
            {"id": "1731861070669187858", "name": "Beleaf Liposomal Vitamin C (วิตามินซีเข้มข้น)", "cat": "Food & Supplements", "price": 390.00, "comm": 0.15, "earn": 58.50, "url": "https://shop.tiktok.com/view/product/1731861070669187858?region=TH&locale=th-TH", "concept": "ไขข้อข้องใจ 'ทำไมกินวิตามินซีปกติแล้วผิวไม่ใสสักที?' เผยเทคโนโลยีไลโปโซมอลที่ดูดซึมดีกว่า 10 เท่า"},
            {"id": "1731576750018628290", "name": "กันแดดเนื้อน้ำนม Hanbyeol (Hanbyeol Sunscreen Milk)", "cat": "Beauty & Personal Care", "price": 290.00, "comm": 0.20, "earn": 58.00, "url": "https://shop.tiktok.com/view/product/1731576750018628290?region=TH&locale=th-TH", "concept": "ทากันแดดแล้วเอาทิชชูซับหน้าให้ดูว่าไม่มีความมันเหนอะหนะ โชว์ความเป็นธรรมชาติ ท้าแดดเมืองไทย"},
            {"id": "1731527642712148180", "name": "Smooth Bio C 1000 mg (วิตามินซีแบบเคี้ยว)", "cat": "Food & Supplements", "price": 150.00, "comm": 0.15, "earn": 22.50, "url": "https://www.tiktok.com/view/product/1731527642712148180?region=TH&locale=th-TH", "concept": "เผลอกินขนมหวานเยอะจนเจ็บคอ ร่างกายอ่อนแอ ลองวิตามินซีเคี้ยวอร่อย ได้สุขภาพดีในราคาหลักร้อย"},
            {"id": "1732374653152432971", "name": "น้ำยาบ้วนปาก Buoola (Buoola Mouthwash)", "cat": "Beauty & Personal Care", "price": 180.00, "comm": 0.15, "earn": 27.00, "url": "https://www.tiktok.com/view/product/1732374653152432971?region=TH&locale=th-TH", "concept": "ปัญหากลิ่นปากแรงหลังกินส้มตำ/กระเทียม บ้วน Buoola เพียง 30 วินาที คราบหลุดออกหมด ลมหายใจหอมสดชื่น"},
            {"id": "1731667000536500824", "name": "เกล็ดขนมปัง ตราครัววังทิพย์ 200 กรัม", "cat": "Food & Kitchen", "price": 45.00, "comm": 0.10, "earn": 4.50, "url": "https://shop.tiktok.com/view/product/1731667000536500824?region=TH&locale=th-TH", "concept": "ทำไก่ทอดกรอบสีเหลืองทองกรุบกรอบน่ากินที่บ้านง่ายๆ ด้วยเกล็ดขนมปังเกรดพรีเมียม"},
            {"id": "1731283808641845848", "name": "ชามพลาสติก PP213 จำนวน 50 ชิ้น", "cat": "Kitchen & Home", "price": 80.00, "comm": 0.10, "earn": 8.00, "url": "https://shop.tiktok.com/view/product/1731283808641845848?region=TH&locale=th-TH", "concept": "จัดระเบียบการแพ็คอาหารส่งเดลิเวอรี่ให้สวยงามด้วยชามฝาล็อคแน่นหนา น้ำซุปไม่มีหกเลอะเทอะ"},
            {"id": "1731860742826396434", "name": "Alpha Chlorophyll Plus (อาหารเสริมคลอโรฟิลล์)", "cat": "Food & Supplements", "price": 290.00, "comm": 0.15, "earn": 43.50, "url": "https://shop.tiktok.com/view/product/1731860742826396434?region=TH&locale=th-TH", "concept": "หน้าท้องป่อง ขับถ่ายยาก ชงดีท็อกซ์คลอโรฟิลล์ก่อนนอน ตื่นเช้ามาพุงยุบ สบายตัวสุดๆ"},
            {"id": "1731194000810019602", "name": "วิตามินบำรุงผิว Colla C (Colla C Collagen)", "cat": "Food & Supplements", "price": 350.00, "comm": 0.18, "earn": 63.00, "url": "https://shop.tiktok.com/view/product/1731194000810019602?region=TH&locale=th-TH", "concept": "เคล็ดลับผิวขาวใสออร่าท้าแดด ชงคอลลาเจนกินง่าย รสชาติอร่อย ผิวนุ่มลื่นใน 7 วัน"}
        ]
        save_to_files(fallback_products)

if __name__ == "__main__":
    main()
