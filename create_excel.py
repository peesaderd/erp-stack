import os
import sys
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
    
    excel_file = "trending_shopee_products.xlsx"
    print(f"Creating highly styled Shopee Excel file: {excel_file}...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shopee Affiliate Products"
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Define styles
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="E24A1C") # Shopee Orange Color
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="333333")
    link_font = Font(name=font_family, size=10, color="2980B9", underline="single")
    
    header_fill = PatternFill(start_color="EE4D2D", end_color="EE4D2D", fill_type="solid") # Shopee Orange
    zebra_fill = PatternFill(start_color="FFF5F2", end_color="FFF5F2", fill_type="solid") # Light Orange-Pink Zebra
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E5E5'),
        right=Side(style='thin', color='E5E5E5'),
        top=Side(style='thin', color='E5E5E5'),
        bottom=Side(style='thin', color='E5E5E5')
    )
    
    # Add title row
    ws.merge_cells("A1:I1")
    ws["A1"] = "สินค้าขายดีอันดับยอดนิยมสำหรับทำวิดีโอ AI (Shopee Affiliate Thailand)"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Add an empty row
    ws.row_dimensions[2].height = 15
    
    # Add headers
    headers = [
        "อันดับ (Rank)",
        "รหัสสินค้า (Product ID)",
        "ชื่อสินค้า (Product Name)",
        "หมวดหมู่ (Category)",
        "ราคาขาย (Price)",
        "ค่าคอมมิชชัน (%)",
        "รายได้ต่อชิ้น (Est. Earnings)",
        "ลิงก์สินค้า (Product URL)",
        "แนวคิด/บทพูดสั้น (AI Video Concept)"
    ]
    
    ws.append([]) # dummy row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[3].height = 30
    
    # Shopee Product Data
    products = [
        {
            "rank": 1, "id": "SHP-MIZUMI-001", "name": "MizuMi UV Water Active Sport SPF50+ PA++++ (40g)",
            "cat": "Beauty & Personal Care", "price": 390.00, "comm": 0.15, "earn": 58.50,
            "url": "https://shopee.co.th/search?keyword=MizuMi%20UV%20Water%20Active%20Sport",
            "concept": "ทากันแดด MizuMi แล้วเอาน้ำสาดหน้าโชว์ความกันน้ำสะเทือนทราย ไม่มีหลุด ซับมันผิวเนียนเรียบ (ผลลัพธ์ใน 3 วินาทีแรก)"
        },
        {
            "rank": 2, "id": "SHP-SIMPLUS-002", "name": "เครื่องดูดฝุ่นไร้สาย Simplus Cordless Vacuum Cleaner (รุ่น XCQH002)",
            "cat": "Home Appliances", "price": 799.00, "comm": 0.12, "earn": 95.88,
            "url": "https://shopee.co.th/search?keyword=Simplus%20Cordless%20Vacuum%20Cleaner",
            "concept": "คลิปดูดเม็ดทรายหรือฝอยเหล็กขนาดเล็กบนเบาะหรือพื้นรถเสียงกร๊อบๆ โชว์พลังดูดแรงในราคาหลักร้อย"
        },
        {
            "rank": 3, "id": "SHP-DRPONG-003", "name": "Dr.PONG 28D Whitening Drone Serum (16ml)",
            "cat": "Beauty & Personal Care", "price": 399.00, "comm": 0.15, "earn": 59.85,
            "url": "https://shopee.co.th/search?keyword=Dr.PONG%2028D%20Whitening%20Drone%20Serum",
            "concept": "โชว์หน้ากระจ่างใสส่องออร่า แล้วอธิบายกลไก Whitening Drone ที่พาเซรั่มลงลึกแก้ไขจุดด่างดำเห็นผลจริง"
        },
        {
            "rank": 4, "id": "SHP-SIMPLUS-004", "name": "หม้อไฟฟ้าอเนกประสงค์ Simplus Multi Cooker 1.5L",
            "cat": "Kitchen Appliances", "price": 299.00, "comm": 0.12, "earn": 35.88,
            "url": "https://shopee.co.th/search?keyword=Simplus%20Multi%20Cooker",
            "concept": "ชงต้มมาม่าใส่เครื่องแน่นๆ ในคอนโด/หอพักเสร็จใน 30 วินาทีแบบรวดเร็ว ล้างหม้อง่าย ไม่ติดกระทะ"
        },
        {
            "rank": 5, "id": "SHP-CLEANER-005", "name": "สเปรย์ขจัดคราบไหม้ก้นกระทะและห้องครัว (Kitchen Cleaner Spray)",
            "cat": "Home Essentials", "price": 99.00, "comm": 0.15, "earn": 14.85,
            "url": "https://shopee.co.th/search?keyword=สเปรย์ขจัดคราบไหม้ก้นกระทะ",
            "concept": "ฉีดสเปรย์ทิ้งไว้ที่ก้นกระทะไหม้ดำปี๋ 10 วินาที แล้วใช้ผ้าเช็ดคราบหลุดออกหมดในรูดเดียว"
        },
        {
            "rank": 6, "id": "SHP-LENOVO-006", "name": "หูฟังไร้สายบลูทูธ Lenovo LP40 Pro",
            "cat": "IT & Gadgets", "price": 250.00, "comm": 0.10, "earn": 25.00,
            "url": "https://shopee.co.th/search?keyword=Lenovo%20LP40%20Pro",
            "concept": "ทดสอบฟังก์ชันแตะเปลี่ยนเพลงและวัดระดับเบสหนักๆ คุ้มเกินราคาหูฟังหลักร้อย สะบัดหัวยังไงก็ไม่หลุด"
        },
        {
            "rank": 7, "id": "SHP-KUKU-007", "name": "กระดาษทิชชู่เปียก Kuku Duckbill Baby Wipes (80 แผ่น)",
            "cat": "Baby & Kids", "price": 45.00, "comm": 0.10, "earn": 4.50,
            "url": "https://shopee.co.th/search?keyword=Kuku%20Duckbill%20Baby%20Wipes",
            "concept": "บีบน้ำชุ่มจากกระดาษทิชชูให้ดูความชุ่มชื้น ไร้กลิ่น ไร้แอลกอฮอล์ เช็ดเล่น เช็ดหน้าเด็ก อ่อนโยนที่สุด"
        },
        {
            "rank": 8, "id": "SHP-STORAGE-008", "name": "กล่องจัดระเบียบตู้เสื้อผ้าแบบลิ้นชักซ้อนแถว",
            "cat": "Home Storage", "price": 120.00, "comm": 0.12, "earn": 14.40,
            "url": "https://shopee.co.th/search?keyword=กล่องจัดระเบียบตู้เสื้อผ้า",
            "concept": "ก่อน-หลังพับเสื้อผ้ากองโตยัดใส่กล่องจัดระเบียบ ตู้เสื้อผ้าโล่ง สะอาดตาหาง่ายในไม่กี่วินาที"
        },
        {
            "rank": 9, "id": "SHP-WANGTHIP-009", "name": "เกล็ดขนมปัง ตราครัววังทิพย์ (200 กรัม)",
            "cat": "Food & Kitchen", "price": 45.00, "comm": 0.08, "earn": 3.60,
            "url": "https://shopee.co.th/search?keyword=เกล็ดขนมปัง%20ครัววังทิพย์",
            "concept": "ชุบแป้งทอดไก่เสียงกรอบฟันหักสีเหลืองทองน่าทานมากด้วยเกล็ดขนมปังครัววังทิพย์"
        },
        {
            "rank": 10, "id": "SHP-TRIPOD-010", "name": "ขาตั้งกล้องบลูทูธ 3-in-1 Selfie Tripod",
            "cat": "IT & Gadgets", "price": 190.00, "comm": 0.10, "earn": 19.00,
            "url": "https://shopee.co.th/search?keyword=ขาตั้งกล้องบลูทูธ",
            "concept": "กางขาตั้งกล้องกว้างๆ ลมพัดก็ไม่ล้ม พร้อมสไลด์ถ่ายแบบกลุ่มสะดวกด้วยรีโมทไร้สายระยะ 10 เมตร"
        }
    ]
    
    current_row = 4
    for p in products:
        ws.cell(row=current_row, column=1, value=p["rank"])
        ws.cell(row=current_row, column=2, value=p["id"])
        ws.cell(row=current_row, column=3, value=p["name"])
        ws.cell(row=current_row, column=4, value=p["cat"])
        ws.cell(row=current_row, column=5, value=p["price"])
        ws.cell(row=current_row, column=6, value=p["comm"])
        ws.cell(row=current_row, column=7, value=p["earn"])
        
        # Hyperlink
        url_cell = ws.cell(row=current_row, column=8, value="Open Link")
        url_cell.hyperlink = p["url"]
        url_cell.font = link_font
        
        ws.cell(row=current_row, column=9, value=p["concept"])
        
        is_even = (current_row % 2 == 0)
        for col_idx in range(1, 10):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            if col_idx != 8:
                cell.font = data_font
            
            if col_idx in [1, 2]:
                cell.alignment = align_center
            elif col_idx in [3, 4, 9]:
                cell.alignment = align_left
            elif col_idx in [5, 6, 7]:
                cell.alignment = align_right
            elif col_idx == 8:
                cell.alignment = align_center
                
            if col_idx in [5, 7]:
                cell.number_format = '฿#,##0.00'
            elif col_idx == 6:
                cell.number_format = '0.0%'
                
            if is_even:
                cell.fill = zebra_fill
                
        ws.row_dimensions[current_row].height = 45
        current_row += 1
        
    # Auto-adjust column widths
    column_widths = {
        "A": 15, "B": 25, "C": 40, "D": 25, "E": 18, "F": 18, "G": 20, "H": 18, "I": 75
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    wb.save(excel_file)
    print(f"Successfully updated Excel with Shopee products: {excel_file}")
    
    # Save CSV version also
    csv_file = "trending_shopee_products.csv"
    with open(csv_file, mode='w', encoding='utf-8-sig', newline='') as f:
        fieldnames = ["Product ID", "Product Name", "Category", "Price (THB)", "Commission Rate", "Product URL", "Google Sheets Link", "AI Video Hook / Concept"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for p in products:
            writer.writerow({
                "Product ID": p["id"],
                "Product Name": p["name"],
                "Category": p["cat"],
                "Price (THB)": f"{p['price']:.2f}",
                "Commission Rate": f"{int(p['comm']*100)}%",
                "Product URL": p["url"],
                "Google Sheets Link": f'=HYPERLINK("{p["url"]}", "Open Link")',
                "AI Video Hook / Concept": p["concept"]
            })
    print(f"Successfully updated CSV with Shopee products: {csv_file}")

if __name__ == "__main__":
    main()
